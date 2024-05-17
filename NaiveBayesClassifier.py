import openpyxl
import csv
import sys
import os
import math
import random
import pandas as pd
import json
import seaborn as sns
import matplotlib.pyplot as plt
import numpy as np

# Function to load data from an Excel file
def load_excel_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    return data

def load_data_from_folder(folder_path):
    all_data = []
    file_names = os.listdir(folder_path)
    xlsx_files = [file_name for file_name in file_names if file_name.endswith('.xlsx')]
    file_numbers = []
    for file_name in xlsx_files:
        numeric_part = ''.join(c for c in file_name if c.isdigit())
        if numeric_part:
            file_numbers.append(int(numeric_part))
        else:
            file_numbers.append(float('inf'))
    file_data = zip(xlsx_files, file_numbers)
    sorted_file_data = sorted(file_data, key=lambda x: x[1])
    for file_name, _ in sorted_file_data:
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(folder_path, file_name)
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            data = [list(row) for row in sheet.iter_rows(values_only=True)]
            all_data.append(data)
    return all_data
# Function to split data into features and target
def preprocess_data(data):
    # Separate features (X) and target (y)
    data = data[1:]
    X = [row[:5] + row[6:] for row in data]  
    y = [row[5] for row in data]    
    return X, y
# Function to calculate the mean
def mean(numbers):
    return sum(numbers) / float(len(numbers))

# Function to calculate the standard deviation
def stdev(numbers):
    avg = mean(numbers)
    if len(numbers) <= 1:
        return 0
    variance = sum([((x - avg) ** 2) for x in numbers]) / float(len(numbers) - 1)
    return math.sqrt(variance)


def calculate_conditional_prob(X, y, feature, feature_value, given_class):
    categorical_features = [0, 4, 5, 6, 7, 9]
    if feature in categorical_features:
        count_c = 1 # to avoid 0 probability (laplace smoothing)
        count_value_given_c = 1
        for i in range(len(X) - 1):  
            if y[i+1] == given_class: 
                count_c += 1
                if X[i][feature] == feature_value:
                    count_value_given_c += 1
        result = count_value_given_c / count_c  
        return result
    return 0.01
    """
    else:
        find_stats = []
        for i in range(len(X) - 1):  
            if y[i+1] == given_class: 
                find_stats.append(X[i][feature])
        mean = np.mean(find_stats)
        std = np.std(find_stats)
        if std == 0: 
            if feature_value == mean:
                return 1.0
            else:
                return 0.0
        else:
            exponent = np.exp(-((feature_value - mean) ** 2) / (2 * std ** 2))
            result = (1 / (np.sqrt(2 * np.pi) * std)) * exponent
            return result
    """

# Function to summarize dataset
def summarize_dataset(X, y):
    # calculate conditional and class probabilities
    class_values = {}
    categorical_features = [0, 4, 5, 6, 7, 9]

    for class_value in y:
        if class_value not in class_values:
            class_values[class_value] = 0
        class_values[class_value] += 1

    for i in class_values:
        class_values[i] = class_values[i] / len(y)

    # conditional values = list of 2d dictionaries, feature/specific feature/class
    conditional_values = [{} for _ in range(len(X[0]))]
    for i in range(len(y)):
        for j in range(len(X[0])):
            feature_value = X[i][j]
            if(feature_value not in conditional_values[j]):
                conditional_values[j][feature_value] = {}
                for key, value in class_values.items():
                    conditional_values[j][feature_value][key] = calculate_conditional_prob(X, y, j, feature_value, key)
    return class_values, conditional_values

# Function to evaluate the model
def evaluate_model(testing, class_probabilities, conditional_probabilities):
    predictions = []
    for data_set in testing:
        test_data = data_set[-1]
        test_data = test_data[:5] + test_data[6:]
        pred = {} 
        for key, value in class_probabilities.items():
            pred[key] = math.log(value) # use log to avoid underflow
            for feature in range(len(test_data)):
                get_value = conditional_probabilities[feature].get(test_data[feature], {}).get(key, 1e-10)
                if(get_value > 0):
                    prob = math.log(get_value)
                    pred[key] += prob
            #print(pred[key], key, value)
        y_pred = max(pred, key=pred.get)
        predictions.append(y_pred)
        print(y_pred)
    return predictions
    

# Main function
def main(training_file, test_folder):
    # Load and preprocess training data
    data = load_excel_data(training_file)
    testing = load_data_from_folder(test_folder)
    X_train, y_train = preprocess_data(data)

    # Train the model
    class_probabilities, conditional_probabilities = summarize_dataset(X_train, y_train)    
    # Evaluate the model on validation data
    #print("Validation Set Evaluation:")
    results = evaluate_model(testing, class_probabilities, conditional_probabilities)
    with open('ground_truth.json', 'r') as file:
        ground_truth = json.load(file)
    correct_predictions = sum(1 for i in range(len(ground_truth)) if ground_truth[i] == results[i])
    accuracy = correct_predictions / len(ground_truth) * 100
   
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python NaiveBayesClassifier.py <training_file> <test_folder>")
        sys.exit(1)
    training_file = sys.argv[1]
    test_folder = sys.argv[2]
    main(training_file, test_folder)
